"""
DCF Excel Template Generator
Generates a single-sheet DCF valuation model from a configuration dictionary.
All calculations use Excel formulas. Run recalc.py after generating.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as cl


def _calc_ic(cfg, idx):
    """Calculate Invested Capital for a given year index.
    IC = Net Working Capital + Net PP&E + Goodwill & Intangibles
    NWC = (CA - Cash - ST Inv - Op Cash) - (CL - ST Debt - ST Leases)
    Shared across all tabs to avoid duplication.
    """
    n = len(cfg['ic_years'])
    ca = cfg['current_assets'][idx]
    cash = cfg['cash'][idx]
    sti = cfg['st_investments'][idx]
    ocash = cfg.get('operating_cash', [0]*n)[idx]
    cl_val = cfg['current_liabilities'][idx]
    std = cfg['st_debt'][idx]
    stl = cfg['st_leases'][idx]
    ppe = cfg['net_ppe'][idx]
    gi = cfg['goodwill_intang'][idx]
    nwc = (ca - cash - sti - ocash) - (cl_val - std - stl)
    return nwc + ppe + gi


def _calc_nopats(cfg):
    """Calculate NOPAT for all historical years."""
    hist_oi = cfg.get('hist_operating_income', [])
    n = len(cfg['ic_years'])
    if len(hist_oi) == n:
        return [oi * (1 - cfg['tax_rate']) for oi in hist_oi]
    return []


def build_dcf_model(cfg, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{cfg.get('ticker','DCF')} Valuation"

    # Styles
    BLUE = Font(name='Calibri', color='0000FF', size=11)
    BLUE_B = Font(name='Calibri', color='0000FF', size=11, bold=True)
    BLK = Font(name='Calibri', color='000000', size=11)
    BLK_B = Font(name='Calibri', color='000000', size=11, bold=True)
    WHT_B = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
    ITAL = Font(name='Calibri', color='000000', size=11, italic=True)
    GRN = Font(name='Calibri', color='008000', size=11)
    HDR = PatternFill('solid', fgColor='4472C4')
    INP = PatternFill('solid', fgColor='BDD7EE')
    THIN = Border(
        left=Side('thin', color='B4C6E7'), right=Side('thin', color='B4C6E7'),
        top=Side('thin', color='B4C6E7'), bottom=Side('thin', color='B4C6E7'))
    CTR = Alignment(horizontal='center')

    n_hist = len(cfg['ic_years'])
    n_proj = len(cfg['revenue_growth'])
    IC = 3  # start column for data
    YR0 = 3; YR1 = 4; YRLAST = YR0 + n_proj; TV = YRLAST + 1

    def sc(r, c, v, f=BLK, fi=None, nf=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.alignment = CTR; cell.border = THIN
        if fi: cell.fill = fi
        if nf: cell.number_format = nf
        return cell

    def lb(r, c, v, f=BLK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.border = THIN
        return cell

    def hdr_bar(r, cs, ce, text=None):
        for c in range(cs, ce + 1):
            ws.cell(row=r, column=c).fill = HDR
            ws.cell(row=r, column=c).font = WHT_B
            ws.cell(row=r, column=c).border = THIN
        if text:
            ws.cell(row=r, column=cs, value=text)

    def note(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = GRN
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        return cell

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 34
    for i in range(3, TV + 2):
        ws.column_dimensions[cl(i)].width = 13
    ws.column_dimensions[cl(TV)].width = 16
    note_col = TV + 1
    ws.column_dimensions[cl(note_col)].width = 45

    # ================================================================
    # SECTION 1: WACC
    # ================================================================
    ws.merge_cells('B2:E2'); hdr_bar(2, 2, 5, f"WACC Calculation — {cfg['company']}")
    ws.merge_cells('H2:J2'); hdr_bar(2, 8, 10, "Levered Beta Calculation")

    lb(4, 2, "Equity Value (Market)"); sc(4, 3, cfg['equity_market_value'], BLUE, None, '#,##0')
    lb(5, 2, "Debt Value (Market)"); sc(5, 3, cfg['debt_market_value'], BLUE, None, '#,##0')
    lb(7, 2, "Risk-Free Rate"); sc(7, 3, cfg['risk_free_rate'], BLUE, None, '0.00%')
    lb(8, 2, "Equity Risk Premium (ERP)"); sc(8, 3, cfg['erp'], BLUE, None, '0.00%')
    lb(9, 2, "Levered Beta"); sc(9, 3, '=J13', BLK, None, '0.00')
    lb(10, 2, "Cost of Equity (CAPM)", BLK_B); sc(10, 3, '=C7+C9*C8', BLK_B, INP, '0.00%')

    lb(12, 2, "Pre-Tax Cost of Debt"); sc(12, 3, '=C7+C13', BLK, None, '0.00%')
    lb(13, 2, "Credit Spread"); sc(13, 3, cfg['credit_spread'], BLUE, None, '0.00%')
    lb(14, 2, "Tax Rate"); sc(14, 3, cfg['tax_rate'], BLUE, None, '0.00%')
    lb(15, 2, "After-Tax Cost of Debt", BLK_B); sc(15, 3, '=C12*(1-C14)', BLK_B, INP, '0.00%')

    lb(17, 2, "E / (D+E)"); sc(17, 3, '=C4/(C4+C5)', BLK, None, '0.00%')
    lb(18, 2, "D / (D+E)"); sc(18, 3, '=C5/(C4+C5)', BLK, None, '0.00%')
    RWACC = 20
    lb(RWACC, 2, "WACC", BLK_B); sc(RWACC, 3, '=C17*C10+C18*C15', BLK_B, INP, '0.00%')

    # Beta calculation — weighted sectors
    betas = cfg['sector_betas']
    for i, (name, ub, wt) in enumerate(betas):
        lb(4 + i, 8, f"Unlevered Beta ({name})"); sc(4 + i, 10, ub, BLUE, None, '0.00')

    if len(betas) > 1:
        r_wt_start = 4 + len(betas)
        for i, (name, ub, wt) in enumerate(betas):
            lb(r_wt_start + i, 8, f"Revenue Weight ({name})"); sc(r_wt_start + i, 10, wt, BLUE, None, '0.0%')

        r_wb = r_wt_start + len(betas)
        lb(r_wb, 8, "Weighted Unlevered Beta", BLK_B)
        parts = '+'.join([f'J{4+i}*J{r_wt_start+i}' for i in range(len(betas))])
        sc(r_wb, 10, f'={parts}', BLK_B, INP, '0.00')
    else:
        r_wb = 4
        lb(r_wb, 8, "Unlevered Beta", BLK_B)

    r_de = r_wb + 2
    lb(r_de, 8, "Equity Value (Market)"); sc(r_de, 10, '=C4', BLK, None, '#,##0')
    lb(r_de + 1, 8, "Debt Value (Market)"); sc(r_de + 1, 10, '=C5', BLK, None, '#,##0')
    lb(r_de + 2, 8, "Tax Rate"); sc(r_de + 2, 10, '=C14', BLK, None, '0.00%')
    lb(r_de + 3, 8, "D/E Ratio"); sc(r_de + 3, 10, f'=J{r_de+1}/J{r_de}', BLK, None, '0.00%')
    R_LB = r_de + 4
    lb(R_LB, 8, "Levered Beta", BLK_B)
    # Row 13 in the original is now R_LB
    # Fix: levered beta formula references weighted unlevered beta
    sc(R_LB, 10, f'=J{r_wb}*(1+(1-J{r_de+2})*J{r_de+3})', BLK_B, INP, '0.00')
    # Update the WACC section reference to levered beta
    ws.cell(row=9, column=3).value = f'=J{R_LB}'

    # Debt breakdown
    r_db = R_LB + 2
    ws.merge_cells(f'H{r_db}:J{r_db}'); hdr_bar(r_db, 8, 10, "Debt Breakdown")
    for i, (lbl, val) in enumerate(cfg['debt_breakdown']):
        lb(r_db + 1 + i, 8, lbl); sc(r_db + 1 + i, 10, val, BLUE, None, '#,##0')
    r_td = r_db + 1 + len(cfg['debt_breakdown'])
    lb(r_td, 8, "Total Debt Value", BLK_B)
    sc(r_td, 10, f'=SUM(J{r_db+1}:J{r_td-1})', BLK_B, INP, '#,##0')
    R_TOTAL_DEBT = r_td

    # ================================================================
    # SECTION 2: INVESTED CAPITAL
    # ================================================================
    R_IC = max(RWACC, r_td) + 4
    ws.merge_cells(f'B{R_IC}:H{R_IC}'); hdr_bar(R_IC, 2, IC + n_hist - 1, "Invested Capital & Sales-to-Capital Ratio")

    r = R_IC + 1
    for i, yr in enumerate(cfg['ic_years']):
        sc(r, IC + i, str(yr), WHT_B, HDR, '@')

    # Helper to add a row of historical data
    def hist_row(row, label, data, font=BLUE, fmt='#,##0'):
        lb(row, 2, label)
        for i, v in enumerate(data):
            sc(row, IC + i, v, font, None, fmt)
        return row

    r = R_IC + 2
    RCA = hist_row(r, "Current Assets", cfg['current_assets']); r += 1
    RCSH = hist_row(r, "  Less: Cash & Equivalents", cfg['cash']); r += 1
    RSTI = hist_row(r, "  Less: ST Investments", cfg['st_investments']); r += 1
    ROPC = hist_row(r, "  Plus: Operating Cash", cfg['operating_cash']); r += 1

    ROCA = r
    lb(r, 2, "Operating Current Assets", BLK_B)
    for i in range(n_hist):
        c = cl(IC + i)
        sc(r, IC + i, f'={c}{RCA}-{c}{RCSH}-{c}{RSTI}+{c}{ROPC}', BLK_B, None, '#,##0')

    r += 2
    RCL = hist_row(r, "Current Liabilities", cfg['current_liabilities']); r += 1
    RSTD = hist_row(r, "  Less: Short-Term Debt", cfg['st_debt']); r += 1
    RSTL = hist_row(r, "  Less: Short-Term Leases", cfg['st_leases']); r += 1

    ROCL = r
    lb(r, 2, "Operating Current Liabilities", BLK_B)
    for i in range(n_hist):
        c = cl(IC + i)
        sc(r, IC + i, f'={c}{RCL}-{c}{RSTD}-{c}{RSTL}', BLK_B, None, '#,##0')

    r += 2
    RNWC = r
    lb(r, 2, "Net Working Capital", BLK_B)
    for i in range(n_hist):
        c = cl(IC + i)
        sc(r, IC + i, f'={c}{ROCA}-{c}{ROCL}', BLK_B, None, '#,##0')

    r += 1; RPPE = hist_row(r, "Net PP&E", cfg['net_ppe'])
    r += 1; RGI = hist_row(r, "Goodwill & Intangibles", cfg['goodwill_intang'])

    r += 2; RICV = r
    lb(r, 2, "Invested Capital", BLK_B)
    for i in range(n_hist):
        c = cl(IC + i)
        sc(r, IC + i, f'={c}{RNWC}+{c}{RPPE}+{c}{RGI}', BLK_B, INP, '#,##0')

    r += 2; RRVH = hist_row(r, "Revenue", cfg['hist_revenue'])

    r += 1; RSTCH = r
    lb(r, 2, "Sales-to-Capital Ratio", BLK_B)
    for i in range(n_hist):
        c = cl(IC + i)
        sc(r, IC + i, f'={c}{RRVH}/{c}{RICV}', BLK_B, None, '0.00')

    r += 1
    lb(r, 2, "Average", BLK_B)
    sc(r, IC, f'=AVERAGE({cl(IC)}{RSTCH}:{cl(IC+n_hist-1)}{RSTCH})', BLK_B, INP, '0.00')

    # ================================================================
    # SECTION 3: DCF
    # ================================================================
    DR = r + 3
    ws.merge_cells(f'B{DR}:{cl(TV)}{DR}')
    hdr_bar(DR, 2, TV, f"Discounted Cash Flows — {cfg['company']}")

    # Year headers
    r = DR + 1
    for i in range(n_proj + 1):
        yr = cfg['base_year'] + i
        sc(r, YR0 + i, str(yr), WHT_B, HDR, '@')
    sc(r, TV, "Terminal Value", ITAL)

    # Mid-year periods
    r = DR + 2
    lb(r, 2, "In millions", BLK_B)
    ynums = [0] + [0.5 + i for i in range(n_proj)]
    for i, yn in enumerate(ynums):
        sc(r, YR0 + i, yn, BLK, None, '0.0' if yn else '0')
    sc(r, TV, ynums[-1], BLK, None, '0.0')

    # Revenue Growth
    RGR = DR + 3
    lb(RGR, 2, "Revenue Growth")
    for i, g in enumerate(cfg['revenue_growth']):
        sc(RGR, YR1 + i, g, BLUE, None, '0.0%')
    sc(RGR, TV, cfg['terminal_growth'], BLUE, None, '0.0%')

    # Revenue
    RREV = DR + 4
    lb(RREV, 2, "Revenue")
    sc(RREV, YR0, cfg['base_revenue'], BLUE, None, '#,##0')
    for i in range(n_proj):
        sc(RREV, YR1 + i, f'={cl(YR0+i)}{RREV}*(1+{cl(YR1+i)}{RGR})', BLK, None, '#,##0')

    # Operating Margin
    RMRG = DR + 5
    lb(RMRG, 2, "Operating Margin")
    sc(RMRG, YR0, cfg['base_op_margin'], BLUE, None, '0.0%')
    for i, m in enumerate(cfg['op_margins']):
        sc(RMRG, YR1 + i, m, BLUE, None, '0.0%')
    sc(RMRG, TV, cfg['terminal_margin'], BLUE, None, '0.0%')

    # Operating Income
    ROI = DR + 6
    lb(ROI, 2, "Operating Income")
    sc(ROI, YR0, cfg['base_oi'], BLUE, None, '#,##0')
    for i in range(n_proj):
        sc(ROI, YR1 + i, f'={cl(YR1+i)}{RREV}*{cl(YR1+i)}{RMRG}', BLK, None, '#,##0')
    sc(ROI, TV, f'={cl(YRLAST)}{RREV}*(1+{cl(TV)}{RGR})*{cl(TV)}{RMRG}', BLK, None, '#,##0')

    # Tax Rate
    RTAX = DR + 7
    lb(RTAX, 2, "Tax Rate")
    for i in range(n_proj + 1):
        sc(RTAX, YR0 + i, f'=$C$14', BLK, None, '0.0%')
    sc(RTAX, TV, f'=$C$14', BLK, None, '0.0%')

    # NOPAT
    RNOP = DR + 8
    lb(RNOP, 2, "NOPAT")
    for i in range(n_proj + 1):
        c = YR0 + i
        sc(RNOP, c, f'={cl(c)}{ROI}*(1-{cl(c)}{RTAX})', BLK, None, '#,##0')
    sc(RNOP, TV, f'={cl(TV)}{ROI}*(1-{cl(TV)}{RTAX})', BLK, None, '#,##0')

    # Sales-to-Capital
    RSTC = DR + 9
    lb(RSTC, 2, "Sales-to-Capital")
    for i in range(n_proj):
        sc(RSTC, YR1 + i, cfg['sales_to_capital'], BLUE, None, '0.0')
    sc(RSTC, TV, cfg['sales_to_capital'], BLUE, None, '0.0')

    # Reinvestment
    RRI = DR + 10
    lb(RRI, 2, "Reinvestment")
    for i in range(n_proj):
        sc(RRI, YR1 + i, f'=({cl(YR1+i)}{RREV}-{cl(YR0+i)}{RREV})/{cl(YR1+i)}{RSTC}', BLK, None, '#,##0')
    sc(RRI, TV, f'=({cl(YRLAST)}{RREV}*(1+{cl(TV)}{RGR})-{cl(YRLAST)}{RREV})/{cl(TV)}{RSTC}', BLK, None, '#,##0')

    # SBC (after tax)
    RSBC = DR + 11
    lb(RSBC, 2, "SBC (after tax)")
    sbc_pct = cfg.get('sbc_pct', 0.004)
    for i in range(n_proj):
        sc(RSBC, YR1 + i, f'={cl(YR1+i)}{RREV}*{sbc_pct}*(1-{cl(YR1+i)}{RTAX})', BLK, None, '#,##0')
    sc(RSBC, TV, f'={cl(YRLAST)}{RREV}*(1+{cl(TV)}{RGR})*{sbc_pct}*(1-{cl(TV)}{RTAX})', BLK, None, '#,##0')

    # FCFF
    RFCFF = DR + 12
    lb(RFCFF, 2, "FCFF", BLK_B)
    for i in range(n_proj):
        c = YR1 + i
        sc(RFCFF, c, f'={cl(c)}{RNOP}-{cl(c)}{RRI}-{cl(c)}{RSBC}', BLK_B, None, '#,##0')
    sc(RFCFF, TV, f'={cl(TV)}{RNOP}-{cl(TV)}{RRI}-{cl(TV)}{RSBC}', BLK_B, None, '#,##0')

    # Undiscounted Terminal Value
    RUTV = DR + 13
    lb(RUTV, 2, "Undiscounted TV")
    RWACC_DCF = DR + 14
    sc(RUTV, TV, f'={cl(TV)}{RFCFF}/({cl(TV)}{RWACC_DCF}-{cl(TV)}{RGR})', BLK, None, '#,##0')

    # WACC row
    RWR = DR + 14
    lb(RWR, 2, "WACC")
    for i in range(n_proj):
        sc(RWR, YR1 + i, f'=$C${RWACC}', BLK, None, '0.00%')
    sc(RWR, TV, f'=$C${RWACC}', BLK, None, '0.00%')

    # Cumulative Discount Factor
    RDF = DR + 15
    lb(RDF, 2, "Cumulative Discount Factor")
    sc(RDF, YR0, 1, BLK, None, '0.00')
    for i in range(n_proj):
        sc(RDF, YR1 + i, f'=1/(1+{cl(YR1+i)}{RWR})^{ynums[i+1]}', BLK, None, '0.0000')
    sc(RDF, TV, f'=1/(1+{cl(TV)}{RWR})^{ynums[-1]}', BLK, None, '0.0000')

    # PV of FCFF
    RPV = DR + 17
    lb(RPV, 2, "PV of FCFF", BLK_B)
    for i in range(n_proj):
        c = YR1 + i
        sc(RPV, c, f'={cl(c)}{RFCFF}*{cl(c)}{RDF}', BLK, None, '#,##0')
    # PV of Terminal Value (CRITICAL: discounted!)
    sc(RPV, TV, f'={cl(TV)}{RUTV}*{cl(TV)}{RDF}', BLK_B, None, '#,##0')
    note(RPV, note_col, "PV of TV = Undiscounted TV × Discount Factor")

    # Enterprise Value
    REV = DR + 19
    lb(REV, 2, "Enterprise Value", BLK_B)
    sc(REV, YR0, f'=SUM({cl(YR1)}{RPV}:{cl(TV)}{RPV})', BLK_B, INP, '#,##0')

    # TV % of EV
    lb(DR + 20, 2, "TV as % of EV")
    sc(DR + 20, YR0, f'={cl(TV)}{RPV}/{cl(YR0)}{REV}', BLK, None, '0.0%')

    # ================================================================
    # SECTION 4: EV TO EQUITY
    # ================================================================
    r = DR + 22
    ws.merge_cells(f'B{r}:C{r}'); hdr_bar(r, 2, 3, "Enterprise Value to Equity Value")

    R_EV2 = DR + 24; lb(R_EV2, 2, "Enterprise Value"); sc(R_EV2, YR0, f'={cl(YR0)}{REV}', BLK, None, '#,##0')
    R_CSH = R_EV2 + 1; lb(R_CSH, 2, "Cash"); sc(R_CSH, YR0, cfg['cash_bridge'], BLUE_B, None, '#,##0')
    R_SEC = R_CSH + 1; lb(R_SEC, 2, "Marketable Securities"); sc(R_SEC, YR0, cfg['securities'], BLUE_B, None, '#,##0')
    R_DBT = R_SEC + 1; lb(R_DBT, 2, "Debt Value"); sc(R_DBT, YR0, f'=J{R_TOTAL_DEBT}', BLK, None, '#,##0')
    R_EQV = R_DBT + 1
    lb(R_EQV, 2, "Equity Value", BLK_B)
    sc(R_EQV, YR0, f'={cl(YR0)}{R_EV2}+{cl(YR0)}{R_CSH}+{cl(YR0)}{R_SEC}-{cl(YR0)}{R_DBT}', BLK_B, None, '#,##0')

    R_SHR_BASE = R_EQV + 2
    lb(R_SHR_BASE, 2, "Shares Outstanding (current)")
    sc(R_SHR_BASE, YR0, cfg['shares_outstanding'], BLUE, None, '#,##0')

    R_BUY_RT = R_SHR_BASE + 1
    lb(R_BUY_RT, 2, "Annual Buyback Reduction")
    sc(R_BUY_RT, YR0, cfg['buyback_rate'], BLUE, None, '0.0%')

    R_SHR_ADJ = R_BUY_RT + 1
    lb(R_SHR_ADJ, 2, f"Adj. Shares Outstanding ({cfg['base_year']+n_proj})", BLK_B)
    sc(R_SHR_ADJ, YR0, f'={cl(YR0)}{R_SHR_BASE}*(1-{cl(YR0)}{R_BUY_RT})^{n_proj}', BLK_B, INP, '#,##0')

    R_PRC = R_SHR_ADJ + 2
    lb(R_PRC, 2, "Share Price", BLK_B)
    sc(R_PRC, YR0, f'={cl(YR0)}{R_EQV}/{cl(YR0)}{R_SHR_ADJ}', BLK_B, INP, '$#,##0.00')

    R_MOS = R_PRC + 1
    lb(R_MOS, 2, "Margin of Safety")
    sc(R_MOS, YR0, cfg['margin_of_safety'], BLUE, None, '0.0%')

    R_BUY = R_MOS + 2
    lb(R_BUY, 2, "Buy Price", BLK_B)
    sc(R_BUY, YR0, f'={cl(YR0)}{R_PRC}*(1-{cl(YR0)}{R_MOS})', BLK_B, INP, '$#,##0.00')

    R_DATE = R_BUY + 1
    lb(R_DATE, 2, "Date"); sc(R_DATE, YR0, cfg.get('valuation_date', ''), BLUE)

    # ================================================================
    # SECTION 5: DYNAMIC SCENARIO ANALYSIS (Formula-Based)
    # ================================================================
    # Pre-compute WACC for reverse DCF / sensitivity (Python-side)
    eq_val = cfg['equity_market_value']
    debt_val = cfg['debt_market_value']
    eq_wt = eq_val / (eq_val + debt_val)
    debt_wt = debt_val / (eq_val + debt_val)
    wu_beta = sum(ub * wt for _, ub, wt in cfg['sector_betas'])
    de_ratio = debt_val / eq_val if eq_val > 0 else 0
    lev_beta = wu_beta * (1 + (1 - cfg['tax_rate']) * de_ratio)
    ke = cfg['risk_free_rate'] + lev_beta * cfg['erp']
    kd = (cfg['risk_free_rate'] + cfg['credit_spread']) * (1 - cfg['tax_rate'])
    cfg['_wacc'] = eq_wt * ke + debt_wt * kd

    def _run_dcf_scenario(cfg, growth_rates, margins):
        """Run a full DCF with given growth/margin assumptions. Returns share price."""
        wacc = cfg['_wacc']
        n_p = len(growth_rates)
        base_rev = cfg['base_revenue']
        tax_r = cfg['tax_rate']
        stc = cfg['sales_to_capital']
        sbc_p = cfg.get('sbc_pct', 0.004)
        tg = cfg['terminal_growth']
        tm = cfg.get('terminal_margin', margins[-1])
        revs = [base_rev]
        for g in growth_rates:
            revs.append(revs[-1] * (1 + g))
        pv_fcff = 0
        for i in range(1, n_p + 1):
            oi = revs[i] * margins[i-1]
            nopat = oi * (1 - tax_r)
            reinvest = (revs[i] - revs[i-1]) / stc
            sbc = revs[i] * sbc_p * (1 - tax_r)
            fcff = nopat - reinvest - sbc
            period = 0.5 + (i - 1)
            df = 1 / (1 + wacc) ** period
            pv_fcff += fcff * df
        tv_rev = revs[-1] * (1 + tg)
        tv_oi = tv_rev * tm
        tv_nopat = tv_oi * (1 - tax_r)
        tv_reinvest = (tv_rev - revs[-1]) / stc
        tv_sbc = tv_rev * sbc_p * (1 - tax_r)
        tv_fcff = tv_nopat - tv_reinvest - tv_sbc
        tv = tv_fcff / (wacc - tg)
        tv_df = 1 / (1 + wacc) ** (0.5 + n_p - 1)
        pv_tv = tv * tv_df
        ev = pv_fcff + pv_tv
        equity = ev + cfg['cash_bridge'] + cfg.get('securities', 0) - cfg['debt_market_value']
        adj_shares = cfg['shares_outstanding'] * (1 - cfg['buyback_rate']) ** n_p
        return equity / adj_shares if adj_shares > 0 else 0

    base_growth = cfg['revenue_growth']
    base_margins = cfg['op_margins']
    mkt_price = cfg.get('stock_price', 0)

    def cagr_from_list(rates):
        prod = 1
        for g in rates: prod *= (1 + g)
        return prod ** (1/len(rates)) - 1

    # --- Scenario Configuration Area ---
    BULL_HDR = PatternFill('solid', fgColor='375623')
    BEAR_HDR = PatternFill('solid', fgColor='953735')
    BULL_BG = PatternFill('solid', fgColor='E2EFDA')
    BEAR_BG = PatternFill('solid', fgColor='FBE5D6')
    BULL_FONT = Font(name='Calibri', color='375623', size=12, bold=True)
    BEAR_FONT = Font(name='Calibri', color='953735', size=12, bold=True)

    bull_g_adj = cfg.get('bull_growth_adj', 0.02)
    bull_m_adj = cfg.get('bull_margin_adj', 0.02)
    bear_g_adj = cfg.get('bear_growth_adj', -0.04)
    bear_m_adj = cfg.get('bear_margin_adj', -0.02)

    R_SC = R_DATE + 2
    lb(R_SC, 2, "Scenario Configuratie", Font(name='Calibri', color='2F5496', size=13, bold=True))

    r = R_SC + 1
    for c, v, fi in [(2, "Aanpassing t.o.v. Base Case", HDR), (4, "Bull", BULL_HDR), (5, "Bear", BEAR_HDR)]:
        sc(r, c, v, WHT_B, fi)
    sc(r, 3, None, WHT_B, HDR)

    r = R_SC + 2
    lb(r, 2, "Revenue Growth Adj (pp)", BLK_B)
    sc(r, 4, bull_g_adj, BLUE_B, None, '+0.0%;-0.0%')
    sc(r, 5, bear_g_adj, BLUE_B, None, '+0.0%;-0.0%')
    R_GADJ = r  # row with growth adjustments

    r = R_SC + 3
    lb(r, 2, "Operating Margin Adj (pp)", BLK_B)
    sc(r, 4, bull_m_adj, BLUE_B, None, '+0.0%;-0.0%')
    sc(r, 5, bear_m_adj, BLUE_B, None, '+0.0%;-0.0%')
    R_MADJ = r  # row with margin adjustments

    r = R_SC + 4
    note(r, 2, "→ Wijzig de blauwe cellen (D/E) om Bull/Bear aan te passen. Alles herberekent automatisch.")

    # --- Helper to build a scenario DCF block ---
    def build_scenario_block(start_row, name, g_adj_cell, m_adj_cell, hdr_fi, bg_fi, sec_font):
        """Build complete formula-linked DCF block. Returns dict with key row refs."""
        r = start_row
        yr_cols = [cl(YR1 + i) for i in range(n_proj)]  # D,E,F,...,M

        # Title
        lb(r, 2, f"{name} Case — Discounted Cash Flows", sec_font)

        # Year headers
        r += 1
        lb(r, 2, "In millions", WHT_B); ws.cell(row=r, column=2).fill = hdr_fi
        for i in range(n_proj + 1):
            c = YR0 + i
            sc(r, c, f'={cl(c)}{DR+1}', WHT_B, hdr_fi)  # ref base year headers
        sc(r, TV, f'={cl(TV)}{DR+1}', WHT_B, hdr_fi)

        # Revenue Growth
        r += 1; rg = r
        lb(r, 2, "Revenue Growth", BLK_B)
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'=MAX({cl(c)}{RGR}+{g_adj_cell},0)', BLK, bg_fi, '0.0%')
        sc(r, TV, f'=MAX({cl(TV)}{RGR}+{g_adj_cell},0)', BLK, bg_fi, '0.0%')

        # Revenue
        r += 1; rev = r
        lb(r, 2, "Revenue", BLK_B)
        sc(r, YR0, f'={cl(YR0)}{RREV}', BLK, None, '#,##0')  # base year = same
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'={cl(c-1)}{rev}*(1+{cl(c)}{rg})', BLK, None, '#,##0')

        # Operating Margin
        r += 1; om = r
        lb(r, 2, "Operating Margin", BLK_B)
        sc(r, YR0, f'={cl(YR0)}{RMRG}', BLK, None, '0.0%')
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'={cl(c)}{RMRG}+{m_adj_cell}', BLK, bg_fi, '0.0%')
        sc(r, TV, f'={cl(TV)}{RMRG}+{m_adj_cell}', BLK, bg_fi, '0.0%')

        # Operating Income
        r += 1; oi = r
        lb(r, 2, "Operating Income", BLK_B)
        sc(r, YR0, f'={cl(YR0)}{rev}*{cl(YR0)}{om}', BLK, None, '#,##0')
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'={cl(c)}{rev}*{cl(c)}{om}', BLK, None, '#,##0')
        sc(r, TV, f'={cl(YRLAST)}{rev}*(1+{cl(TV)}{rg})*{cl(TV)}{om}', BLK, None, '#,##0')

        # Tax Rate
        r += 1; tx = r
        lb(r, 2, "Tax Rate")
        for c in range(YR0, TV + 1):
            sc(r, c, f'=$C$14', BLK, None, '0.0%')

        # NOPAT
        r += 1; nop = r
        lb(r, 2, "NOPAT", BLK_B)
        for c in range(YR0, TV + 1):
            sc(r, c, f'={cl(c)}{oi}*(1-{cl(c)}{tx})', BLK, None, '#,##0')

        # Sales-to-Capital (ref base)
        r += 1; stc = r
        lb(r, 2, "Sales-to-Capital")
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'={cl(c)}{RSTC}', BLK, None, '0.0')
        sc(r, TV, f'={cl(TV)}{RSTC}', BLK, None, '0.0')

        # Reinvestment
        r += 1; reinv = r
        lb(r, 2, "Reinvestment")
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'=({cl(c)}{rev}-{cl(c-1)}{rev})/{cl(c)}{stc}', BLK, None, '#,##0')
        sc(r, TV, f'=({cl(YRLAST)}{rev}*(1+{cl(TV)}{rg})-{cl(YRLAST)}{rev})/{cl(TV)}{stc}', BLK, None, '#,##0')

        # SBC
        r += 1; sbc_r = r
        lb(r, 2, "SBC (after tax)")
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'={cl(c)}{rev}*{sbc_pct}*(1-{cl(c)}{tx})', BLK, None, '#,##0')
        sc(r, TV, f'={cl(YRLAST)}{rev}*(1+{cl(TV)}{rg})*{sbc_pct}*(1-{cl(TV)}{tx})', BLK, None, '#,##0')

        # FCFF
        r += 1; fcff = r
        lb(r, 2, "FCFF", BLK_B)
        for c in range(YR1, TV + 1):
            sc(r, c, f'={cl(c)}{nop}-{cl(c)}{reinv}-{cl(c)}{sbc_r}', BLK_B, None, '#,##0')

        # WACC (ref base)
        r += 1; wacc_r = r
        lb(r, 2, "WACC")
        for c in range(YR1, TV + 1):
            sc(r, c, f'=$C${RWACC}', BLK, None, '0.00%')

        # Undiscounted TV
        r += 1; tv_r = r
        lb(r, 2, "Undiscounted TV")
        sc(r, TV, f'={cl(TV)}{fcff}/({cl(TV)}{wacc_r}-{cl(TV)}{rg})', BLK, None, '#,##0')

        # Discount Factor
        r += 1; df = r
        lb(r, 2, "Discount Factor")
        sc(r, YR0, 1, BLK, None, '0.00')
        for i in range(n_proj):
            c = YR1 + i
            period = 0.5 + i
            sc(r, c, f'=1/(1+{cl(c)}{wacc_r})^{period}', BLK, None, '0.0000')
        sc(r, TV, f'=1/(1+{cl(TV)}{wacc_r})^{ynums[-1]}', BLK, None, '0.0000')

        # PV of FCFF
        r += 2; pv = r
        lb(r, 2, "PV of FCFF", BLK_B)
        for i in range(n_proj):
            c = YR1 + i
            sc(r, c, f'={cl(c)}{fcff}*{cl(c)}{df}', BLK, None, '#,##0')
        sc(r, TV, f'={cl(TV)}{tv_r}*{cl(TV)}{df}', BLK_B, None, '#,##0')

        # Enterprise Value
        r += 2; ev_r = r
        lb(r, 2, "Enterprise Value", BLK_B)
        sc(r, YR0, f'=SUM({cl(YR1)}{pv}:{cl(TV)}{pv})', BLK_B, INP, '#,##0')

        # Equity Value
        r += 1; eq_r = r
        lb(r, 2, "Equity Value", BLK_B)
        sc(r, YR0, f'={cl(YR0)}{ev_r}+{cl(YR0)}{R_CSH}+{cl(YR0)}{R_SEC}-J{R_TOTAL_DEBT}', BLK, None, '#,##0')

        # Share Price
        r += 1; sp = r
        lb(r, 2, "Share Price", BLK_B)
        sc(r, YR0, f'={cl(YR0)}{eq_r}/{cl(YR0)}{R_SHR_ADJ}', BLK_B, bg_fi, '$#,##0.00')
        ws.cell(row=r, column=YR0).font = Font(name='Calibri', size=13, bold=True)

        # vs Market
        r += 1
        lb(r, 2, "vs Marktprijs", BLK_B)
        if mkt_price:
            sc(r, YR0, f'=({cl(YR0)}{sp}/{mkt_price})-1', BLK, None, '+0.0%;-0.0%')

        # Buy Price
        r += 1
        lb(r, 2, "Buy Price (incl MoS)", BLK_B)
        sc(r, YR0, f'={cl(YR0)}{sp}*(1-{cl(YR0)}{R_MOS})', BLK, None, '$#,##0.00')

        return {'end': r, 'price': sp, 'rev': rev, 'om': om, 'rg': rg}

    # --- Build Bull Case ---
    BULL_START = R_SC + 6
    bull = build_scenario_block(BULL_START, "Bull", f"$D${R_GADJ}", f"$D${R_MADJ}",
                                 BULL_HDR, BULL_BG, BULL_FONT)

    # --- Build Bear Case ---
    BEAR_START = bull['end'] + 3
    bear = build_scenario_block(BEAR_START, "Bear", f"$E${R_GADJ}", f"$E${R_MADJ}",
                                 BEAR_HDR, BEAR_BG, BEAR_FONT)

    # --- Scenario Summary Table ---
    r = bear['end'] + 3
    lb(r, 2, "Scenario Overzicht", Font(name='Calibri', color='2F5496', size=13, bold=True))
    r += 1
    for c, v in [(2,"Scenario"),(3,"Share Price"),(4,"Buy Price"),(5,"vs Markt"),(6,"Revenue CAGR"),(7,"Gem. Marge")]:
        sc(r, c, v, WHT_B, HDR)
    R_SUM_HDR = r

    # Bull summary row
    r += 1
    lb(r, 2, "Bull", Font(name='Calibri', color='375623', bold=True))
    ws.cell(row=r, column=2).fill = BULL_BG
    sc(r, 3, f'={cl(YR0)}{bull["price"]}', BLK, BULL_BG, '$#,##0.00')
    sc(r, 4, f'={cl(YR0)}{bull["price"]}*(1-{cl(YR0)}{R_MOS})', BLK, None, '$#,##0.00')
    if mkt_price:
        sc(r, 5, f'=({cl(YR0)}{bull["price"]}/{mkt_price})-1', BLK, None, '+0.0%;-0.0%')
    sc(r, 6, f'=({cl(YRLAST)}{bull["rev"]}/{cl(YR0)}{bull["rev"]})^(1/{n_proj})-1', BLK, None, '0.0%')
    sc(r, 7, f'=AVERAGE({cl(YR1)}{bull["om"]}:{cl(YRLAST)}{bull["om"]})', BLK, None, '0.0%')

    # Base summary row
    r += 1
    lb(r, 2, "Base", Font(name='Calibri', color='2F5496', bold=True))
    ws.cell(row=r, column=2).fill = INP
    sc(r, 3, f'={cl(YR0)}{R_PRC}', BLK, INP, '$#,##0.00')
    sc(r, 4, f'={cl(YR0)}{R_BUY}', BLK, None, '$#,##0.00')
    if mkt_price:
        sc(r, 5, f'=({cl(YR0)}{R_PRC}/{mkt_price})-1', BLK, None, '+0.0%;-0.0%')
    sc(r, 6, f'=({cl(YRLAST)}{RREV}/{cl(YR0)}{RREV})^(1/{n_proj})-1', BLK, None, '0.0%')
    sc(r, 7, f'=AVERAGE({cl(YR1)}{RMRG}:{cl(YRLAST)}{RMRG})', BLK, None, '0.0%')

    # Bear summary row
    r += 1
    lb(r, 2, "Bear", Font(name='Calibri', color='953735', bold=True))
    ws.cell(row=r, column=2).fill = BEAR_BG
    sc(r, 3, f'={cl(YR0)}{bear["price"]}', BLK, BEAR_BG, '$#,##0.00')
    sc(r, 4, f'={cl(YR0)}{bear["price"]}*(1-{cl(YR0)}{R_MOS})', BLK, None, '$#,##0.00')
    if mkt_price:
        sc(r, 5, f'=({cl(YR0)}{bear["price"]}/{mkt_price})-1', BLK, None, '+0.0%;-0.0%')
    sc(r, 6, f'=({cl(YRLAST)}{bear["rev"]}/{cl(YR0)}{bear["rev"]})^(1/{n_proj})-1', BLK, None, '0.0%')
    sc(r, 7, f'=AVERAGE({cl(YR1)}{bear["om"]}:{cl(YRLAST)}{bear["om"]})', BLK, None, '0.0%')

    # Market price reference
    r += 1
    lb(r, 2, "Marktprijs", BLK_B)
    sc(r, 3, mkt_price, BLK, None, '$#,##0.00')

    r += 2
    note(r, 2, "Bull/Bear refereren met formules aan de Base Case + adjustments. Wijzig Base (rij "
         f"{RGR}-{RSTC}) of adjustments (rij {R_GADJ}-{R_MADJ}) en alles updatet automatisch.")
    
    # ================================================================
    # SECTION 6: REVERSE DCF — 2D Sensitivity Matrix
    # ================================================================
    R_RD = r + 2
    
    # Growth rates (rows) and margin levels (columns)
    growth_tests = [0.02, 0.04, 0.06, 0.08, 0.10, 0.12, 0.15, 0.18, 0.20, 0.25]
    
    # Margin tests: center around base avg margin, ±6% range in 2% steps
    avg_base_margin = sum(base_margins) / len(base_margins)
    margin_center = round(avg_base_margin * 20) / 20  # round to nearest 5%
    margin_tests = sorted(set([
        max(0.10, margin_center - 0.06),
        max(0.10, margin_center - 0.04),
        max(0.10, margin_center - 0.02),
        margin_center,
        min(0.50, margin_center + 0.02),
        min(0.50, margin_center + 0.04),
        min(0.50, margin_center + 0.06),
    ]))
    
    n_grow = len(growth_tests)
    n_marg = len(margin_tests)
    
    # Header
    matrix_width = 2 + n_marg  # label col + margin cols
    ws.merge_cells(f'B{R_RD}:{cl(2 + n_marg + 1)}{R_RD}')
    hdr_bar(R_RD, 2, 2 + n_marg + 1, "Reverse DCF — Revenue Growth × Operating Margin Sensitivity")
    
    r = R_RD + 1
    note(r, 2, f"Implied share price at each growth/margin combination | WACC: {cfg['_wacc']:.2%} | Market: ${mkt_price:.2f}")
    
    # Column headers (margins)
    r += 1
    lb(r, 2, "CAGR \\ Margin", BLK_B)
    for j, m in enumerate(margin_tests):
        is_base = abs(m - avg_base_margin) < 0.005
        font = BLK_B if is_base else BLK
        sc(r, 3 + j, m, font, HDR if is_base else None, '0.0%')
        if is_base:
            ws.cell(row=r, column=3 + j).font = WHT_B
    
    # Pre-compute all prices & find closest to market
    matrix = {}
    closest_cell = None
    closest_diff = float('inf')
    
    for rate in growth_tests:
        for margin in margin_tests:
            uniform_margins = [margin] * len(base_margins)
            uniform_growth = [rate] * len(base_growth)
            price = _run_dcf_scenario(cfg, uniform_growth, uniform_margins)
            matrix[(rate, margin)] = price
            diff = abs(price - mkt_price)
            if diff < closest_diff:
                closest_diff = diff
                closest_cell = (rate, margin)
    
    # Fill matrix rows
    for i, rate in enumerate(growth_tests):
        r += 1
        is_base_growth = abs(rate - cagr_from_list(base_growth)) < 0.015
        font_label = BLK_B if is_base_growth else BLK
        sc(r, 2, rate, font_label, HDR if is_base_growth else None, '0.0%')
        if is_base_growth:
            ws.cell(row=r, column=2).font = WHT_B
        
        for j, margin in enumerate(margin_tests):
            price = matrix[(rate, margin)]
            
            # Color: green if above market (undervalued), red if below (overvalued)
            is_closest = (rate, margin) == closest_cell
            is_base_combo = is_base_growth and abs(margin - avg_base_margin) < 0.005
            
            if is_closest:
                # Highlight market-implied cell
                cell_font = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
                cell_fill = PatternFill('solid', fgColor='FF6600')  # orange
            elif is_base_combo:
                cell_font = BLK_B
                cell_fill = INP
            elif price >= mkt_price:
                cell_font = Font(name='Calibri', color='008000', size=11)
                cell_fill = PatternFill('solid', fgColor='E2EFDA')  # light green
            else:
                cell_font = Font(name='Calibri', color='CC0000', size=11)
                cell_fill = None
            
            sc(r, 3 + j, round(price, 2), cell_font, cell_fill, '$#,##0')
    
    # Legend
    r += 2
    lb(r, 2, "Legend", BLK_B)
    cell = ws.cell(row=r, column=3, value="■")
    cell.font = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
    cell.fill = PatternFill('solid', fgColor='FF6600')
    cell.border = THIN
    lb(r, 4, "Closest to market price (what market implies)", BLK)
    
    r += 1
    cell = ws.cell(row=r, column=3, value="■")
    cell.font = Font(name='Calibri', color='008000', size=11)
    cell.fill = PatternFill('solid', fgColor='E2EFDA')
    cell.border = THIN
    lb(r, 4, "Price above market → undervalued at those assumptions", BLK)
    
    r += 1
    cell = ws.cell(row=r, column=3, value="■")
    cell.font = Font(name='Calibri', color='CC0000', size=11)
    cell.border = THIN
    lb(r, 4, "Price below market → overvalued at those assumptions", BLK)
    
    r += 1
    if closest_cell:
        cg, cm = closest_cell
        note(r, 2, f"Market implies ~{cg:.0%} revenue CAGR at {cm:.0%} margin (or equivalent combo) to justify ${mkt_price:.2f}")
    
    r += 1
    base_cagr = cagr_from_list(base_growth)
    base_price_check = matrix.get(closest_cell, 0)
    if closest_cell and closest_cell[0] > base_cagr * 1.2:
        note(r, 2, f"⚠ Market is more optimistic: implies {closest_cell[0]:.0%} growth at {closest_cell[1]:.0%} margin vs your {base_cagr:.1%}/{avg_base_margin:.1%}")
    elif closest_cell and closest_cell[0] < base_cagr * 0.8:
        note(r, 2, f"✓ Potential undervaluation: market only requires {closest_cell[0]:.0%} growth vs your {base_cagr:.1%} base case")
    else:
        note(r, 2, f"≈ Fairly priced: market-implied assumptions are close to your base case")
    
    # Clean up temp key after all sheets are built (moved to after save)

    # Add Summary tab if historical data available
    if cfg.get('hist_operating_income'):
        build_summary_sheet(wb, cfg)
        build_calculations_sheet(wb, cfg)

    # Add Peer Comparison tab if peer data available
    if cfg.get('peers'):
        build_peer_comparison_sheet(wb, cfg)

    # Add Sensitivity Analysis tab
    build_sensitivity_sheet(wb, cfg, _run_dcf_scenario)

    wb.save(output_path)
    # Clean up temp key
    if '_wacc' in cfg:
        del cfg['_wacc']
    print(f"Model saved to {output_path}")
    print(f"Sections: WACC(row 2), IC(row {R_IC}), DCF(row {DR}), Equity(row {DR+22})")
    print(f"Key rows: WACC={RWACC}, EV={REV}, Price={R_PRC}, Buy={R_BUY}")
    print(f"Dynamic scenarios: Bull(row {BULL_START}), Bear(row {BEAR_START})")
    return output_path


def build_summary_sheet(wb, cfg):
    """Add a Summary tab as the first sheet in the workbook.
    
    Required extra config keys (beyond DCF):
        hist_operating_income: list of operating income values matching ic_years
        hist_net_income: list of net income values matching ic_years
        hist_sbc_values: list of SBC values matching ic_years
        hist_shares: list of diluted share counts matching ic_years
        stock_price: current stock price (or derived from equity_market_value/shares_outstanding)
    
    Handles:
        - Asset-light companies (negative IC → N/A for ROIC, FCF ROC)
        - Loss years (negative OI/NI → N/A for growth from negative base)
        - ROIIC computed as 3-year rolling to smooth volatility
    """
    ws = wb.create_sheet("Summary", 0)
    
    # ── Styles ──
    BLUE = Font(name='Calibri', color='0000FF', size=11)
    BLK = Font(name='Calibri', color='000000', size=11)
    BLK_B = Font(name='Calibri', color='000000', size=11, bold=True)
    WHT_B = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
    GRN_I = Font(name='Calibri', color='008000', size=11, italic=True)
    GREY = Font(name='Calibri', color='808080', size=11, italic=True)
    RED = Font(name='Calibri', color='FF0000', size=11)
    HDR = PatternFill('solid', fgColor='4472C4')
    INP = PatternFill('solid', fgColor='BDD7EE')
    QUAL_BG = PatternFill('solid', fgColor='F2F2F2')
    SUB_HDR = PatternFill('solid', fgColor='D6E4F0')
    THIN = Border(
        left=Side('thin', color='B4C6E7'), right=Side('thin', color='B4C6E7'),
        top=Side('thin', color='B4C6E7'), bottom=Side('thin', color='B4C6E7'))
    CTR = Alignment(horizontal='center')
    WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def sc(r, c, v, f=BLK, fi=None, nf=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.alignment = CTR; cell.border = THIN
        if fi: cell.fill = fi
        if nf: cell.number_format = nf
        return cell
    
    def lb(r, c, v, f=BLK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.border = THIN
        return cell
    
    def hdr_bar(r, cs, ce, text=None):
        for c in range(cs, ce + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = HDR; cell.font = WHT_B; cell.border = THIN
        if text:
            ws.cell(row=r, column=cs, value=text)
    
    def note(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = GRN_I; cell.alignment = WRAP
        return cell
    
    def sub_hdr(r, cs, ce, text):
        lb(r, cs, text, BLK_B)
        for c in range(cs, ce + 1):
            ws.cell(row=r, column=c).fill = SUB_HDR
    
    # ── Data setup ──
    years = cfg['ic_years']
    n = len(years)
    D0 = 3
    DLAST = D0 + n - 1
    COL_AVG = DLAST + 1
    COL_TREND = COL_AVG + 1
    RP = COL_TREND + 2  # right panel label
    RV = RP + 1          # right panel value
    
    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    for i in range(D0, COL_TREND + 1):
        ws.column_dimensions[cl(i)].width = 13
    ws.column_dimensions[cl(RP)].width = 28
    ws.column_dimensions[cl(RV)].width = 14
    ws.column_dimensions[cl(RV + 1)].width = 30
    
    # ── Derived data ──
    hist_oi = cfg.get('hist_operating_income', [])
    hist_ni = cfg.get('hist_net_income', [])
    hist_sbc = cfg.get('hist_sbc_values', [])
    hist_shares = cfg.get('hist_shares', [])
    tax = cfg['tax_rate']
    
    ics = [_calc_ic(cfg, i) for i in range(n)]
    nopats = _calc_nopats(cfg)
    
    stock_price = cfg.get('stock_price', cfg['equity_market_value'] / cfg['shares_outstanding'])
    mkt_cap = cfg['equity_market_value']
    ev = mkt_cap + cfg['debt_market_value'] - cfg['cash_bridge'] - cfg.get('securities', 0)
    
    # ════════════════════════════════════════════
    # HEADER
    # ════════════════════════════════════════════
    r = 2
    hdr_bar(r, 2, COL_TREND, f"Company Summary — {cfg['company']} ({cfg['ticker']})")
    
    r = 3
    lb(r, 2, "Date"); sc(r, D0, cfg.get('valuation_date', ''), BLUE)
    lb(r, D0 + 1, "Price"); sc(r, D0 + 2, stock_price, BLUE, None, '$#,##0.00')
    lb(r, D0 + 3, "Market Cap ($M)"); sc(r, D0 + 4, mkt_cap, BLUE, None, '#,##0')
    
    # ════════════════════════════════════════════
    # QUALITATIVE (left) + VALUATION RATIOS (right)
    # ════════════════════════════════════════════
    r_qual = 5
    hdr_bar(r_qual, 2, D0 + 2, "Qualitative Assessment")
    
    qual_items = [
        ("Business Case", "Brief description of what the company does"),
        ("Competitive Moat", "Network effects, brand, switching costs, scale, IP"),
        ("Management & Insiders", "Founder-led? CEO ownership? Insider buying?"),
    ]
    r = r_qual + 1
    for label, hint in qual_items:
        lb(r, 2, label, BLK_B)
        note(r, D0 + 3, hint)
        for rr in range(r + 1, r + 4):
            for cc in range(2, D0 + 3):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = QUAL_BG; cell.border = THIN; cell.alignment = WRAP
        r += 4
    r_qual_end = r
    
    # ── Right: Valuation Ratios ──
    hdr_bar(r_qual, RP, RV + 1, "Valuation Snapshot")
    rv = r_qual + 1
    
    # P/E Ratio
    last_ni = hist_ni[-1] if hist_ni else None
    lb(rv, RP, "P/E Ratio")
    if last_ni and last_ni > 0:
        sc(rv, RV, mkt_cap / last_ni, BLK, None, '0.0x')
    else:
        sc(rv, RV, "N/A (loss)", GREY)
    note(rv, RV + 1, "S&P 500 historical avg: 19.4x")
    
    rv += 1
    lb(rv, RP, "EV / Revenue")
    last_rev = cfg['hist_revenue'][-1]
    sc(rv, RV, ev / last_rev if last_rev else 0, BLK, None, '0.0x')
    note(rv, RV + 1, f"EV = ${ev:,.0f}M")
    
    rv += 1
    last_nopat = nopats[-1] if nopats else 0
    lb(rv, RP, "EV / NOPAT")
    sc(rv, RV, ev / last_nopat if last_nopat and last_nopat > 0 else "N/A", 
       BLK if last_nopat and last_nopat > 0 else GREY, None, '0.0x' if last_nopat and last_nopat > 0 else None)
    
    rv += 1
    # FCF Yield
    fcf_last = None
    if len(nopats) >= 2 and len(ics) >= 2:
        fcf_last = nopats[-1] - (ics[-1] - ics[-2])
    lb(rv, RP, "FCF Yield")
    if fcf_last and mkt_cap > 0:
        sc(rv, RV, fcf_last / mkt_cap, BLK, None, '0.0%')
    else:
        sc(rv, RV, "N/A", GREY)
    note(rv, RV + 1, f"Compare to risk-free: {cfg['risk_free_rate']:.1%}")
    
    rv += 1
    net_cash = cfg['cash_bridge'] + cfg.get('securities', 0) - cfg['debt_market_value']
    lb(rv, RP, "Net Cash per Share")
    sc(rv, RV, net_cash / cfg['shares_outstanding'], BLK, None, '$#,##0.00')
    note(rv, RV + 1, f"{net_cash/mkt_cap*100:.0f}% of market cap" if mkt_cap else "")
    
    rv += 1
    last_sbc = hist_sbc[-1] if hist_sbc else 0
    lb(rv, RP, "SBC / Market Cap")
    if last_sbc and mkt_cap > 0:
        sc(rv, RV, last_sbc / mkt_cap, BLK, None, '0.0%')
    else:
        sc(rv, RV, "N/A", GREY)
    note(rv, RV + 1, "Annual dilution cost to shareholders")
    
    rv += 1
    lb(rv, RP, "Buyback Yield")
    if len(hist_shares) >= 2 and hist_shares[-2] != 0:
        sc(rv, RV, -(hist_shares[-1] - hist_shares[-2]) / hist_shares[-2], BLK, None, '0.0%')
        note(rv, RV + 1, "Positive = net buybacks")
    else:
        sc(rv, RV, "N/A", GREY)
    
    rv += 1
    lb(rv, RP, "PEG Ratio")
    sc(rv, RV, "", BLUE)  # manual input
    note(rv, RV + 1, "P/E ÷ expected growth rate. <1 = attractive")
    
    # ════════════════════════════════════════════
    # HISTORICAL METRICS TABLE
    # ════════════════════════════════════════════
    r_hist = r_qual_end + 1
    hdr_bar(r_hist, 2, COL_TREND)
    ws.cell(row=r_hist, column=2, value="Historical Metrics")
    
    r = r_hist + 1
    lb(r, 2, "In millions", GREY)
    for i, yr in enumerate(years):
        sc(r, D0 + i, yr, BLK_B)
    sc(r, COL_AVG, "Average", BLK_B)
    sc(r, COL_TREND, "Trend", BLK_B)
    
    # ── Metric row helper ──
    def metric_row(row, label, values, fmt='0.0%', note_text=None, font=BLK, na_check=None):
        lb(row, 2, label, font)
        valid_vals = []
        val_cols = []
        for i, v in enumerate(values):
            col = D0 + i
            if v is None or (na_check and na_check(v)):
                sc(row, col, "N/A", GREY)
            else:
                sc(row, col, v, BLUE, None, fmt)
                valid_vals.append(v)
                val_cols.append(cl(col))
        # Average
        if valid_vals and len(valid_vals) >= 2:
            avg_f = f'=AVERAGE({",".join([f"{c}{row}" for c in val_cols])})'
            sc(row, COL_AVG, avg_f, BLK, INP, fmt)
        else:
            sc(row, COL_AVG, "N/A", GREY)
        # Trend
        if valid_vals and len(valid_vals) >= 3:
            if valid_vals[-1] > valid_vals[0]:
                sc(row, COL_TREND, "↑ Improving", Font(name='Calibri', color='008000', size=11))
            elif valid_vals[-1] < valid_vals[0]:
                sc(row, COL_TREND, "↓ Declining", RED)
            else:
                sc(row, COL_TREND, "→ Stable", BLK)
        if note_text:
            note(row, COL_TREND + 1, note_text)
    
    # ── MARGINS ──
    r = r_hist + 2
    sub_hdr(r, 2, COL_TREND, "MARGINS")
    
    r += 1
    op_margins = [hist_oi[i] / cfg['hist_revenue'][i] if cfg['hist_revenue'][i] else None
                  for i in range(n)] if len(hist_oi) == n else [None]*n
    metric_row(r, "Operating Margin", op_margins, note_text="Sector dependent target")
    
    r += 1
    net_margins = [hist_ni[i] / cfg['hist_revenue'][i] if cfg['hist_revenue'][i] else None
                   for i in range(n)] if len(hist_ni) == n else [None]*n
    metric_row(r, "Net Margin", net_margins)
    
    r += 1
    if len(hist_oi) == n and len(hist_sbc) == n:
        sbc_adj = [(hist_oi[i] - hist_sbc[i]) / cfg['hist_revenue'][i]
                   if cfg['hist_revenue'][i] else None for i in range(n)]
    else:
        sbc_adj = [None]*n
    metric_row(r, "SBC-adj Operating Margin", sbc_adj,
               note_text="Op Margin minus SBC — true tech profitability")
    
    r += 1
    fcf_margins = [None]
    if nopats and len(ics) == n:
        for i in range(1, n):
            fcf = nopats[i] - (ics[i] - ics[i-1])
            rev = cfg['hist_revenue'][i]
            fcf_margins.append(fcf / rev if rev else None)
    else:
        fcf_margins = [None]*n
    metric_row(r, "FCF Margin", fcf_margins, note_text="NOPAT − ΔIC / Revenue")
    
    # ── RETURNS ON CAPITAL ──
    r += 2
    sub_hdr(r, 2, COL_TREND, "RETURNS ON CAPITAL")
    
    r += 1
    roics = [nopats[i] / ics[i] if ics[i] > 0 else None for i in range(n)] if nopats else [None]*n
    metric_row(r, "ROIC (NOPAT / IC)", roics, na_check=lambda v: v is None,
               note_text="N/A when IC ≤ 0 (asset-light)")
    
    r += 1
    roiics = [None] * min(3, n)
    if nopats and len(nopats) >= 4:
        for i in range(3, n):
            d_nopat = nopats[i] - nopats[i-3]
            d_ic = ics[i] - ics[i-3]
            roiics.append(d_nopat / d_ic if abs(d_ic) > 10 else None)
    else:
        roiics = [None]*n
    metric_row(r, "ROIIC (3yr rolling)", roiics, na_check=lambda v: v is None,
               note_text="ΔNOPAT / ΔIC over 3 years. N/A if ΔIC ≈ 0")
    
    r += 1
    fcf_rocs = [None]
    if nopats and len(ics) >= 2:
        for i in range(1, n):
            fcf_i = nopats[i] - (ics[i] - ics[i-1])
            fcf_rocs.append(fcf_i / ics[i] if ics[i] > 0 else None)
    else:
        fcf_rocs = [None]*n
    metric_row(r, "FCF Return on Capital", fcf_rocs, na_check=lambda v: v is None,
               note_text="N/A when IC ≤ 0. Target: >15%")
    
    # ── GROWTH ──
    r += 2
    sub_hdr(r, 2, COL_TREND, "GROWTH")
    
    r += 1
    rev = cfg['hist_revenue']
    rev_growth = [None] + [(rev[i]-rev[i-1])/rev[i-1] if rev[i-1] else None for i in range(1, n)]
    metric_row(r, "Revenue Growth", rev_growth)
    
    r += 1
    hist_cogs = cfg.get('hist_cost_of_revenue', [])
    if len(hist_cogs) == n:
        cogs_growth = [None] + [(hist_cogs[i]-hist_cogs[i-1])/hist_cogs[i-1] 
                                 if hist_cogs[i-1] and hist_cogs[i-1] > 0 else None for i in range(1, n)]
    else:
        cogs_growth = [None]*n
    metric_row(r, "Cost of Revenue Growth", cogs_growth, na_check=lambda v: v is None,
               note_text="< Revenue Growth = gross margin expansion")
    
    r += 1
    # Operating Expenses = Revenue - COGS - Operating Income
    if len(hist_cogs) == n and len(hist_oi) == n:
        hist_opex = [rev[i] - hist_cogs[i] - hist_oi[i] for i in range(n)]
        opex_growth = [None] + [(hist_opex[i]-hist_opex[i-1])/hist_opex[i-1]
                                 if hist_opex[i-1] and hist_opex[i-1] > 0 else None for i in range(1, n)]
    else:
        opex_growth = [None]*n
    metric_row(r, "Operating Expenses Growth", opex_growth, na_check=lambda v: v is None,
               note_text="< Revenue Growth = operating leverage")
    
    r += 1
    oi_growth = [None]
    if len(hist_oi) == n:
        for i in range(1, n):
            oi_growth.append((hist_oi[i]-hist_oi[i-1])/hist_oi[i-1] if hist_oi[i-1] > 0 else None)
    else:
        oi_growth = [None]*n
    metric_row(r, "Operating Income Growth", oi_growth, na_check=lambda v: v is None)
    
    r += 1
    ni_growth = [None]
    if len(hist_ni) == n:
        for i in range(1, n):
            ni_growth.append((hist_ni[i]-hist_ni[i-1])/hist_ni[i-1] if hist_ni[i-1] > 0 else None)
    else:
        ni_growth = [None]*n
    metric_row(r, "Net Income Growth", ni_growth, na_check=lambda v: v is None)
    
    r += 1
    share_growth = [None]
    if len(hist_shares) == n:
        for i in range(1, n):
            share_growth.append((hist_shares[i]-hist_shares[i-1])/hist_shares[i-1] if hist_shares[i-1] else None)
    else:
        share_growth = [None]*n
    metric_row(r, "Share Count Growth", share_growth,
               note_text="Negative = buybacks (positive for shareholders)")
    
    # ── FINANCIAL HEALTH ──
    r += 2
    sub_hdr(r, 2, COL_TREND, "FINANCIAL HEALTH")
    
    r += 1
    ca = cfg['current_assets']; cliab = cfg['current_liabilities']
    current_ratios = [ca[i]/cliab[i] if cliab[i] else None for i in range(n)]
    metric_row(r, "Current Ratio", current_ratios, '0.00x',
               note_text=">1.5 healthy, <1.0 watch liquidity")
    
    r += 1
    nd_nopat = []
    for i in range(n):
        cash_total = cfg['cash'][i] + cfg['st_investments'][i]
        total_debt = cfg['st_debt'][i] + cfg['st_leases'][i]
        if i == n - 1:
            total_debt += cfg.get('debt_market_value', 0)
        net_d = total_debt - cash_total
        nd_nopat.append(net_d / nopats[i] if nopats and nopats[i] > 0 else None)
    metric_row(r, "Net Debt / NOPAT", nd_nopat, '0.0x', na_check=lambda v: v is None,
               note_text="<3x healthy. Negative = net cash")
    
    # ── COMPOSITE METRICS ──
    r += 2
    sub_hdr(r, 2, COL_TREND, "COMPOSITE METRICS")
    
    r += 1
    rule40 = [None]
    for i in range(1, n):
        rg = rev_growth[i]
        fm = fcf_margins[i] if i < len(fcf_margins) else None
        rule40.append((rg + fm) * 100 if rg is not None and fm is not None else None)
    metric_row(r, "Rule of 40 Score", rule40, '0.0', na_check=lambda v: v is None,
               note_text="Rev Growth% + FCF Margin%. >40 = strong")
    
    # ── Reference data ──
    r += 1
    if nopats:
        lb(r, 2, "NOPAT ($M)", GREY)
        for i, np_val in enumerate(nopats):
            sc(r, D0 + i, round(np_val), GREY, None, '#,##0')
    
    r += 1
    lb(r, 2, "Invested Capital ($M)", GREY)
    for i, ic_val in enumerate(ics):
        sc(r, D0 + i, round(ic_val), GREY, None, '#,##0')
    
    r += 1
    lb(r, 2, "FCF ($M)", GREY)
    if nopats and len(ics) >= 2:
        sc(r, D0, "—", GREY)
        for i in range(1, n):
            sc(r, D0 + i, round(nopats[i] - (ics[i] - ics[i-1])), GREY, None, '#,##0')
    
    r += 2
    note(r, 2, "FCF = NOPAT − ΔIC (Damodaran method) | N/A when metric not meaningful (negative IC, base-year loss)")
    
    return ws


def build_calculations_sheet(wb, cfg):
    """Add a Calculations tab with FCF/Share analysis and price history structure.
    
    Shows TWO FCF definitions side by side to expose SBC distortion:
    - True FCF/Share: NOPAT − ΔIC (Damodaran) — treats SBC as real cost
    - Screener FCF/Share: True FCF + SBC — what Yahoo/screeners report (inflated)
    
    Auto-fills: FCF/Share (both), SBC/Share, Yield, Growth, CAGR, Expected Return, R²
    Manual: Price history (paste from Google Finance), Price CAGR
    """
    ws = wb.create_sheet("Calculations")
    
    # ── Styles ──
    BLUE = Font(name='Calibri', color='0000FF', size=11)
    BLUE_B = Font(name='Calibri', color='0000FF', size=11, bold=True)
    BLK = Font(name='Calibri', color='000000', size=11)
    BLK_B = Font(name='Calibri', color='000000', size=11, bold=True)
    WHT_B = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
    GRN_I = Font(name='Calibri', color='008000', size=11, italic=True)
    GREY = Font(name='Calibri', color='808080', size=11, italic=True)
    RED = Font(name='Calibri', color='CC0000', size=11)
    RED_B = Font(name='Calibri', color='CC0000', size=11, bold=True)
    HDR = PatternFill('solid', fgColor='4472C4')
    INP = PatternFill('solid', fgColor='BDD7EE')
    WARN = PatternFill('solid', fgColor='FCE4EC')   # light red for SBC distortion
    EMPTY = PatternFill('solid', fgColor='FFF2CC')   # light yellow for manual input
    THIN = Border(
        left=Side('thin', color='B4C6E7'), right=Side('thin', color='B4C6E7'),
        top=Side('thin', color='B4C6E7'), bottom=Side('thin', color='B4C6E7'))
    CTR = Alignment(horizontal='center')
    WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def sc(r, c, v, f=BLK, fi=None, nf=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.alignment = CTR; cell.border = THIN
        if fi: cell.fill = fi
        if nf: cell.number_format = nf
        return cell
    
    def lb(r, c, v, f=BLK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.border = THIN
        return cell
    
    def hdr_bar(r, cs, ce, text=None):
        for c in range(cs, ce + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = HDR; cell.font = WHT_B; cell.border = THIN
        if text:
            ws.cell(row=r, column=cs, value=text)
    
    def note(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = GRN_I; cell.alignment = WRAP
        return cell
    
    # ── Data ──
    years = cfg['ic_years']
    n = len(years)
    tax = cfg['tax_rate']
    hist_oi = cfg.get('hist_operating_income', [])
    hist_shares = cfg.get('hist_shares', [])
    hist_sbc = cfg.get('hist_sbc_values', [])
    stock_price = cfg.get('stock_price', cfg['equity_market_value'] / cfg['shares_outstanding'])
    
    ics = [_calc_ic(cfg, i) for i in range(n)]
    nopats = _calc_nopats(cfg)
    
    # True FCF = NOPAT − ΔIC (SBC treated as real cost)
    true_fcfs = [None]
    if nopats:
        for i in range(1, n):
            true_fcfs.append(nopats[i] - (ics[i] - ics[i-1]))
    
    # SBC per share
    sbc_per_share = []
    if len(hist_sbc) == n and len(hist_shares) == n:
        sbc_per_share = [hist_sbc[i] / hist_shares[i] if hist_shares[i] else 0 for i in range(n)]
    
    # True FCF/Share and Screener FCF/Share
    true_fcf_ps = []
    screener_fcf_ps = []
    for i in range(n):
        if true_fcfs[i] is not None and len(hist_shares) > i and hist_shares[i]:
            t = true_fcfs[i] / hist_shares[i]
            true_fcf_ps.append(t)
            # Screener = True FCF + SBC (add back the non-cash SBC expense)
            s = t + (sbc_per_share[i] if i < len(sbc_per_share) else 0)
            screener_fcf_ps.append(s)
        else:
            true_fcf_ps.append(None)
            screener_fcf_ps.append(None)
    
    # ── Column widths ──
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 16   # True FCF/Share
    ws.column_dimensions['I'].width = 16   # Screener FCF/Share
    ws.column_dimensions['J'].width = 12   # SBC/Share
    ws.column_dimensions['K'].width = 14   # FCF Yield
    ws.column_dimensions['L'].width = 14   # R² Linearity
    ws.column_dimensions['M'].width = 14   # Growth Rate
    ws.column_dimensions['N'].width = 14   # CAGR
    ws.column_dimensions['O'].width = 18   # Expected Return
    
    # ════════════════════════════════════════════
    # LEFT: PRICE HISTORY
    # ════════════════════════════════════════════
    r = 2
    hdr_bar(r, 2, 3, f"Price History — {cfg['company']} ({cfg['ticker']})")
    hdr_bar(r, 7, 15, f"FCF/Share Analysis — {cfg['company']} ({cfg['ticker']})")
    
    r = 3
    lb(r, 2, "Name"); sc(r, 3, cfg['company'], BLK)
    r = 4
    lb(r, 2, "Ticker"); sc(r, 3, cfg['ticker'], BLK)
    r = 5
    lb(r, 2, "Current Price"); sc(r, 3, stock_price, BLUE, None, '$#,##0.00')
    r = 6
    lb(r, 2, "Price 5 years ago"); sc(r, 3, "", BLUE, EMPTY, '$#,##0.00')
    note(r, 5, "Manual / Google Finance")
    r = 7
    lb(r, 2, "Price 10 years ago"); sc(r, 3, "", BLUE, EMPTY, '$#,##0.00')
    r = 8
    lb(r, 2, "Price 20 years ago"); sc(r, 3, "", BLUE, EMPTY, '$#,##0.00')
    
    r = 10
    lb(r, 2, "CAGR 5 years", BLK_B)
    sc(r, 3, '=IF(C6="","",((C5/C6)^(1/5))-1)', BLK, INP, '0.00%')
    r = 11
    lb(r, 2, "CAGR 10 years", BLK_B)
    sc(r, 3, '=IF(C7="","",((C5/C7)^(1/10))-1)', BLK, INP, '0.00%')
    r = 12
    lb(r, 2, "CAGR 20 years", BLK_B)
    sc(r, 3, '=IF(C8="","",((C5/C8)^(1/20))-1)', BLK, INP, '0.00%')
    
    # ════════════════════════════════════════════
    # RIGHT: FCF/SHARE TABLE (dual: True vs Screener)
    # ════════════════════════════════════════════
    
    # Column headers
    r = 3
    sc(r, 7, "Year", BLK_B)
    sc(r, 8, "True FCF/Sh", BLK_B)
    sc(r, 9, "Screener FCF/Sh", RED_B)
    sc(r, 10, "SBC/Share", RED_B)
    sc(r, 11, "True FCF Yield", BLK_B)
    sc(r, 12, "R² Linearity", BLK_B)
    sc(r, 13, "Growth Rate", BLK_B)
    sc(r, 14, "CAGR", BLK_B)
    sc(r, 15, "Exp. Return (5yr)", BLK_B)
    
    # Sub-header notes
    r = 4
    sc(r, 8, "NOPAT−ΔIC", GRN_I); sc(r, 9, "True + SBC", GRN_I)
    sc(r, 10, "SBC distortion", GRN_I); sc(r, 11, "at current price", GRN_I)
    sc(r, 12, "FCF/Sh trend", GRN_I); sc(r, 13, "True FCF YoY", GRN_I)
    sc(r, 14, "True FCF", GRN_I); sc(r, 15, "Yield + CAGR", GRN_I)
    
    # Historical rows
    r_fcf_start = 5
    valid_true_rows = []
    
    for i in range(n):
        r = r_fcf_start + i
        sc(r, 7, years[i], BLK)
        
        # True FCF/Share
        if true_fcf_ps[i] is not None:
            sc(r, 8, round(true_fcf_ps[i], 2), BLUE, None, '#,##0.00')
            valid_true_rows.append(r)
        else:
            sc(r, 8, "N/A", GREY)
        
        # Screener FCF/Share (highlighted to show inflation)
        if screener_fcf_ps[i] is not None:
            sc(r, 9, round(screener_fcf_ps[i], 2), RED, WARN, '#,##0.00')
        else:
            sc(r, 9, "N/A", GREY)
        
        # SBC/Share
        if i < len(sbc_per_share):
            sc(r, 10, round(sbc_per_share[i], 2), RED, WARN, '#,##0.00')
        else:
            sc(r, 10, "", GREY)
        
        # True FCF Yield (only last historical year)
        if i == n - 1 and true_fcf_ps[i] is not None:
            sc(r, 11, f'=H{r}/C5', BLK, None, '0.00%')
        else:
            sc(r, 11, "", GREY)
        
        # Growth rate (True FCF YoY)
        if i > 0 and true_fcf_ps[i] is not None and true_fcf_ps[i-1] is not None:
            if true_fcf_ps[i-1] > 0:
                sc(r, 13, f'=(H{r}-H{r-1})/ABS(H{r-1})', BLK, None, '0.00%')
            else:
                sc(r, 13, "N/A", GREY)
        else:
            sc(r, 13, "", GREY)
    
    r_fcf_hist_end = r_fcf_start + n - 1
    
    # R² Linearity of True FCF/Share (last historical row)
    if len(valid_true_rows) >= 3:
        yr_range = f'G{valid_true_rows[0]}:G{valid_true_rows[-1]}'
        fcf_range = f'H{valid_true_rows[0]}:H{valid_true_rows[-1]}'
        sc(r_fcf_hist_end, 12, f'=RSQ({fcf_range},{yr_range})', BLK, INP, '0.00')
    
    # CAGR of True FCF/Share
    if len(valid_true_rows) >= 2:
        first_r = valid_true_rows[0]
        last_r = valid_true_rows[-1]
        n_yrs = len(valid_true_rows) - 1
        sc(r_fcf_hist_end, 14, 
           f'=IF(H{first_r}>0,(H{last_r}/H{first_r})^(1/{n_yrs})-1,"N/A")',
           BLK, INP, '0.00%')
    
    # Expected 5yr return = True FCF Yield + True FCF CAGR
    sc(r_fcf_hist_end, 15, 
       f'=IF(AND(ISNUMBER(K{r_fcf_hist_end}),ISNUMBER(N{r_fcf_hist_end})),K{r_fcf_hist_end}+N{r_fcf_hist_end},"N/A")',
       BLK_B, INP, '0.00%')
    
    # ── Projected FCF/Share ──
    n_proj = len(cfg['revenue_growth'])
    r_proj_start = r_fcf_hist_end + 1
    
    # Last known SBC ratio to revenue (for projecting SBC forward)
    last_sbc_ratio = cfg.get('sbc_pct', 0)
    
    for j in range(n_proj):
        r = r_proj_start + j
        proj_year = cfg['base_year'] + 1 + j
        sc(r, 7, proj_year, GREY)
        
        # True FCF/Share projection (using revenue growth)
        sc(r, 8, f'=H{r-1}*(1+{cfg["revenue_growth"][j]})', GREY, None, '#,##0.00')
        
        # Screener FCF projection = True + projected SBC/share
        # SBC grows with revenue but we keep the ratio constant
        if last_sbc_ratio > 0:
            sc(r, 9, f'=H{r}+J{r}', GREY, WARN, '#,##0.00')
            # SBC/Share projection: grow from last known
            if j == 0:
                sc(r, 10, f'=J{r-1}*(1+{cfg["revenue_growth"][j]})', GREY, WARN, '#,##0.00')
            else:
                sc(r, 10, f'=J{r-1}*(1+{cfg["revenue_growth"][j]})', GREY, WARN, '#,##0.00')
        else:
            sc(r, 9, f'=H{r}', GREY, None, '#,##0.00')
            sc(r, 10, 0, GREY, None, '#,##0.00')
        
        # FCF Yield at current price
        sc(r, 11, f'=H{r}/C5', GREY, None, '0.00%')
        
        # Growth rate
        sc(r, 13, f'=(H{r}-H{r-1})/ABS(H{r-1})', GREY, None, '0.00%')
    
    r_proj_end = r_proj_start + n_proj - 1
    
    # Projected CAGR
    sc(r_proj_end, 14, 
       f'=IF(H{r_fcf_hist_end}>0,(H{r_proj_end}/H{r_fcf_hist_end})^(1/{n_proj})-1,"N/A")',
       GREY, None, '0.00%')
    
    # Projected R² (historical + projected)
    if valid_true_rows:
        all_yr = f'G{valid_true_rows[0]}:G{r_proj_end}'
        all_fcf = f'H{valid_true_rows[0]}:H{r_proj_end}'
        sc(r_proj_end, 12, f'=RSQ({all_fcf},{all_yr})', GREY, None, '0.00')
    
    # ── Regression stats ──
    r_reg = r_proj_end + 2
    lb(r_reg, 7, "Regression Stats", BLK_B)
    note(r_reg, 8, "True FCF/Share trend")
    
    r_reg += 1
    lb(r_reg, 7, "Slope ($/yr)")
    if len(valid_true_rows) >= 2:
        yr_r = f'G{valid_true_rows[0]}:G{valid_true_rows[-1]}'
        fc_r = f'H{valid_true_rows[0]}:H{valid_true_rows[-1]}'
        sc(r_reg, 8, f'=SLOPE({fc_r},{yr_r})', BLK, INP, '0.000')
    
    r_reg += 1
    lb(r_reg, 7, "Intercept")
    if len(valid_true_rows) >= 2:
        sc(r_reg, 8, f'=INTERCEPT({fc_r},{yr_r})', BLK, INP, '0.000')
    
    # ── SBC & Dilution Analysis ──
    r_sbc = r_reg + 2
    lb(r_sbc, 7, "SBC & Dilution", BLK_B)
    note(r_sbc, 8, "Is SBC a real cost or offset by buybacks?")
    
    # Header row
    r_sbc += 1
    sc(r_sbc, 7, "", BLK_B)
    for i in range(n):
        sc(r_sbc, 8 + i, years[i], BLK_B)
    
    # Row: Gross SBC ($M)
    r_sbc += 1
    lb(r_sbc, 7, "Gross SBC ($M)")
    for i in range(n):
        if i < len(hist_sbc):
            sc(r_sbc, 8 + i, round(hist_sbc[i]), RED, WARN, '#,##0')
        else:
            sc(r_sbc, 8 + i, "", GREY)
    
    # Row: Share Count (M)
    r_sbc += 1
    lb(r_sbc, 7, "Shares (M)")
    for i in range(n):
        if i < len(hist_shares):
            sc(r_sbc, 8 + i, hist_shares[i], BLK, None, '#,##0')
    
    # Row: Net Share Change (%)
    r_sbc += 1
    lb(r_sbc, 7, "Net Dilution %")
    note(r_sbc, 8, "")  # first year = no prior
    for i in range(1, n):
        if hist_shares[i-1] and hist_shares[i-1] > 0:
            chg = (hist_shares[i] - hist_shares[i-1]) / hist_shares[i-1]
            # Green if shares decreasing (buybacks > SBC), red if increasing
            font = GRN_I if chg < 0 else RED
            sc(r_sbc, 8 + i, round(chg, 4), font, None, '0.00%')
    
    # Row: SBC as % of Market Cap (annual dilution cost at current price)
    r_sbc += 1
    lb(r_sbc, 7, "SBC/Market Cap")
    mkt_cap = cfg.get('equity_market_value', stock_price * cfg['shares_outstanding'])
    for i in range(n):
        if i < len(hist_sbc) and mkt_cap > 0:
            sc(r_sbc, 8 + i, round(hist_sbc[i] / mkt_cap, 4), RED, WARN, '0.00%')
    note(r_sbc, 8 + n, "Annual dilution cost at today's price")
    
    # Summary verdict
    r_sbc += 2
    lb(r_sbc, 7, "Verdict", BLK_B)
    
    # Calculate: are shares net declining over the full period?
    if len(hist_shares) >= 2:
        first_shares = hist_shares[0]
        last_shares = hist_shares[-1]
        # Find peak shares (often right after IPO lockup)
        peak_shares = max(hist_shares)
        peak_idx = hist_shares.index(peak_shares)
        total_chg = (last_shares - first_shares) / first_shares
        peak_chg = (last_shares - peak_shares) / peak_shares if peak_shares > 0 else 0
        
        # Recent trend (last 3 years)
        recent_start = max(0, n - 4)
        recent_chg = (hist_shares[-1] - hist_shares[recent_start]) / hist_shares[recent_start]
        
        if recent_chg < -0.02:  # >2% net reduction recently
            verdict = f"Buybacks > SBC: shares down {abs(recent_chg):.1%} since {years[recent_start]} despite ${hist_sbc[-1]}M SBC/yr"
            sc(r_sbc, 8, verdict, GRN_I)
            note(r_sbc + 1, 8, "→ SBC is largely offset. True FCF is conservative, reality is between True and Screener.")
        elif recent_chg > 0.02:  # >2% net dilution recently
            verdict = f"SBC > Buybacks: shares up {recent_chg:.1%} since {years[recent_start]}. SBC is a real cost."
            sc(r_sbc, 8, verdict, RED)
            note(r_sbc + 1, 8, "→ True FCF is the better measure. Screener FCF is misleading.")
        else:
            verdict = f"Roughly neutral: shares ~flat since {years[recent_start]} ({recent_chg:+.1%})"
            sc(r_sbc, 8, verdict, BLK)
            note(r_sbc + 1, 8, "→ SBC approximately offset by buybacks. Both measures are reasonable.")
    
    r_sbc += 3
    
    # ── TTM Stats ──
    r_ttm = r_sbc
    lb(r_ttm, 7, "TTM Stats", BLK_B)
    
    r_ttm += 1
    lb(r_ttm, 7, "Shares (#M)")
    sc(r_ttm, 8, cfg['shares_outstanding'], BLUE, None, '#,##0')
    
    r_ttm += 1
    last_fcf = true_fcfs[-1] if true_fcfs and true_fcfs[-1] is not None else 0
    lb(r_ttm, 7, "True FCF TTM ($M)")
    sc(r_ttm, 8, round(last_fcf), BLK, None, '#,##0')
    
    r_ttm += 1
    lb(r_ttm, 7, "True FCF/Sh TTM")
    sc(r_ttm, 8, f'=H{r_ttm-1}/H{r_ttm-2}', BLK, INP, '#,##0.00')
    
    r_ttm += 1
    lb(r_ttm, 7, "True FCF Yield", BLK_B)
    sc(r_ttm, 8, f'=H{r_ttm-1}/C5', BLK_B, INP, '0.00%')
    
    r_ttm += 1
    lb(r_ttm, 7, "Screener FCF TTM", RED_B)
    screener_last = last_fcf + (hist_sbc[-1] if hist_sbc else 0)
    sc(r_ttm, 8, round(screener_last), RED, WARN, '#,##0')
    
    r_ttm += 1
    lb(r_ttm, 7, "Screener Yield", RED)
    sc(r_ttm, 8, f'=(H{r_ttm-1}/H{r_ttm-5})/C5', RED, WARN, '0.00%')
    
    # ════════════════════════════════════════════
    # WEEKLY PRICE HISTORY (left column, empty structure)
    # ════════════════════════════════════════════
    r_price = 17
    lb(r_price, 2, "Date", BLK_B)
    lb(r_price, 3, "Close", BLK_B)
    ws.cell(row=r_price, column=2).fill = HDR; ws.cell(row=r_price, column=2).font = WHT_B
    ws.cell(row=r_price, column=3).fill = HDR; ws.cell(row=r_price, column=3).font = WHT_B
    
    note(r_price, 5, "Paste weekly close prices from Google Finance / Yahoo Finance")
    
    for i in range(520):
        r = r_price + 1 + i
        cell_d = ws.cell(row=r, column=2)
        cell_d.fill = EMPTY; cell_d.border = THIN; cell_d.number_format = 'MM/DD/YYYY'
        cell_p = ws.cell(row=r, column=3)
        cell_p.fill = EMPTY; cell_p.border = THIN; cell_p.number_format = '$#,##0.00'
    
    return ws


def build_peer_comparison_sheet(wb, cfg):
    """Add a Peer Comparison tab with valuation multiples and quality metrics.

    Required config key:
        peers: list of dicts with keys: ticker, name, ev_revenue, ev_ebitda, pe,
               op_margin, rev_growth, roic
    """
    ws = wb.create_sheet("Peer Comparison", 1)

    # Styles
    BLUE = Font(name='Calibri', color='0000FF', size=11)
    BLK = Font(name='Calibri', color='000000', size=11)
    BLK_B = Font(name='Calibri', color='000000', size=11, bold=True)
    WHT_B = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
    GRN = Font(name='Calibri', color='008000', size=11)
    GRN_I = Font(name='Calibri', color='008000', size=11, italic=True)
    RED = Font(name='Calibri', color='CC0000', size=11)
    GREY = Font(name='Calibri', color='808080', size=11, italic=True)
    HDR = PatternFill('solid', fgColor='4472C4')
    INP = PatternFill('solid', fgColor='BDD7EE')
    GRN_BG = PatternFill('solid', fgColor='E2EFDA')
    RED_BG = PatternFill('solid', fgColor='FBE5D6')
    THIN = Border(
        left=Side('thin', color='B4C6E7'), right=Side('thin', color='B4C6E7'),
        top=Side('thin', color='B4C6E7'), bottom=Side('thin', color='B4C6E7'))
    CTR = Alignment(horizontal='center')

    def sc(r, c, v, f=BLK, fi=None, nf=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.alignment = CTR; cell.border = THIN
        if fi: cell.fill = fi
        if nf: cell.number_format = nf
        return cell

    def lb(r, c, v, f=BLK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.border = THIN
        return cell

    peers = cfg.get('peers', [])
    ticker = cfg.get('ticker', 'Subject')
    company = cfg.get('company', 'Subject')
    n_peers = len(peers)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    for i in range(3, 3 + n_peers + 3):
        ws.column_dimensions[cl(i)].width = 14

    # Header
    r = 2
    for c in range(2, 3 + n_peers + 3):
        ws.cell(row=r, column=c).fill = HDR
        ws.cell(row=r, column=c).font = WHT_B
        ws.cell(row=r, column=c).border = THIN
    ws.cell(row=r, column=2, value=f"Peer Comparison — {company} ({ticker})")

    # Column headers
    r = 3
    lb(r, 2, "Metric", BLK_B)
    sc(r, 3, ticker, WHT_B, HDR)
    for i, p in enumerate(peers):
        sc(r, 4 + i, p.get('ticker', p.get('name', '')), BLK_B)
    col_avg = 4 + n_peers
    col_med = col_avg + 1
    sc(r, col_avg, "Peer Avg", BLK_B, INP)
    sc(r, col_med, "Peer Med", BLK_B, INP)

    # Metrics to display
    mkt_cap = cfg.get('equity_market_value', 0)
    debt = cfg.get('debt_market_value', 0)
    cash = cfg.get('cash_bridge', 0) + cfg.get('securities', 0)
    ev_val = mkt_cap + debt - cash
    last_rev = cfg.get('hist_revenue', [0])[-1]
    last_oi = cfg.get('hist_operating_income', [0])[-1]
    hist_ni = cfg.get('hist_net_income', [0])
    last_ni = hist_ni[-1] if hist_ni else 0
    hist_rev = cfg.get('hist_revenue', [])
    rev_g = (hist_rev[-1] / hist_rev[-2] - 1) if len(hist_rev) >= 2 and hist_rev[-2] else 0
    op_margin = last_oi / last_rev if last_rev else 0
    stock_price = cfg.get('stock_price', 0)

    # Compute subject company EBITDA proxy (OI + D&A estimate)
    # Use OI as proxy if no EBITDA available
    ebitda_est = last_oi * 1.15  # rough: OI + ~15% for D&A
    ev_ebitda_subj = ev_val / ebitda_est if ebitda_est else 0
    ev_rev_subj = ev_val / last_rev if last_rev else 0
    pe_subj = mkt_cap / last_ni if last_ni and last_ni > 0 else None

    # Invested capital for ROIC
    ics = []
    n_hist = len(cfg.get('ic_years', []))
    for i in range(n_hist):
        ics.append(_calc_ic(cfg, i))
    nopats = _calc_nopats(cfg)
    roic_subj = nopats[-1] / ics[-1] if nopats and ics and ics[-1] > 0 else None

    metrics = [
        ("EV / Revenue", ev_rev_subj, 'ev_revenue', '0.0x'),
        ("EV / EBITDA", ev_ebitda_subj, 'ev_ebitda', '0.0x'),
        ("P/E Ratio", pe_subj, 'pe', '0.0x'),
        ("Operating Margin", op_margin, 'op_margin', '0.0%'),
        ("Revenue Growth", rev_g, 'rev_growth', '0.0%'),
        ("ROIC", roic_subj, 'roic', '0.0%'),
    ]

    for label, subj_val, peer_key, fmt in metrics:
        r += 1
        lb(r, 2, label, BLK_B)

        # Subject company
        if subj_val is not None:
            sc(r, 3, subj_val, BLUE, INP, fmt)
        else:
            sc(r, 3, "N/A", GREY, INP)

        # Peer values
        peer_vals = []
        for i, p in enumerate(peers):
            val = p.get(peer_key)
            if val is not None:
                sc(r, 4 + i, val, BLK, None, fmt)
                peer_vals.append(val)
            else:
                sc(r, 4 + i, "N/A", GREY)

        # Average and median
        if peer_vals:
            avg = sum(peer_vals) / len(peer_vals)
            sorted_vals = sorted(peer_vals)
            mid = len(sorted_vals) // 2
            med = sorted_vals[mid] if len(sorted_vals) % 2 else (sorted_vals[mid-1] + sorted_vals[mid]) / 2
            sc(r, col_avg, avg, BLK, INP, fmt)
            sc(r, col_med, med, BLK, INP, fmt)

    # Multiple-based valuation scenarios
    r += 3
    for c in range(2, col_med + 1):
        ws.cell(row=r, column=c).fill = HDR
        ws.cell(row=r, column=c).font = WHT_B
        ws.cell(row=r, column=c).border = THIN
    ws.cell(row=r, column=2, value="Multiple-Based Valuation Scenarios")

    r += 1
    lb(r, 2, "Scenario", BLK_B)
    sc(r, 3, "EV/EBITDA Used", BLK_B)
    sc(r, 4, "Implied EV ($M)", BLK_B)
    sc(r, 5, "Implied Price", BLK_B)
    sc(r, 6, "vs Market", BLK_B)

    # Gather EV/EBITDA values for scenarios
    peer_ebitda_vals = [p.get('ev_ebitda') for p in peers if p.get('ev_ebitda')]
    if peer_ebitda_vals and ebitda_est:
        avg_mult = sum(peer_ebitda_vals) / len(peer_ebitda_vals)
        shares_adj = cfg['shares_outstanding'] * (1 - cfg['buyback_rate']) ** len(cfg['revenue_growth'])

        scenarios = [("At Peer Average", avg_mult)]
        # Add highest and lowest peer
        if len(peers) >= 2:
            sorted_peers = sorted(peers, key=lambda p: p.get('ev_ebitda', 0), reverse=True)
            top_peer = sorted_peers[0]
            bot_peer = sorted_peers[-1]
            scenarios.append((f"At {top_peer.get('ticker','')} Multiple", top_peer.get('ev_ebitda', avg_mult)))
            scenarios.append((f"At {bot_peer.get('ticker','')} Multiple", bot_peer.get('ev_ebitda', avg_mult)))

        for label, mult in scenarios:
            r += 1
            lb(r, 2, label)
            sc(r, 3, mult, BLK, None, '0.0x')
            implied_ev = mult * ebitda_est
            sc(r, 4, round(implied_ev), BLK, None, '#,##0')
            implied_eq = implied_ev - debt + cash
            implied_price = implied_eq / shares_adj if shares_adj else 0
            sc(r, 5, round(implied_price, 2), BLK_B, INP, '$#,##0.00')
            if mkt_price := cfg.get('stock_price', 0):
                upside = (implied_price / mkt_price) - 1
                font = GRN if upside > 0 else RED
                sc(r, 6, upside, font, None, '+0.0%;-0.0%')

    # Verdict
    r += 2
    lb(r, 2, "Beoordeling", BLK_B)
    r += 1
    ws.cell(row=r, column=2, value="DCF vs Peer vergelijking: zie hoe intrinsieke waarde zich verhoudt tot relative waardering.")
    ws.cell(row=r, column=2).font = GRN_I

    return ws


def build_sensitivity_sheet(wb, cfg, run_dcf_fn):
    """Add a Sensitivity Analysis tab with three 2D matrices.

    Matrix 1: Revenue CAGR × Operating Margin (at base WACC)
    Matrix 2: Revenue CAGR × WACC (at base margin)
    Matrix 3: Operating Margin × WACC (at base CAGR)
    """
    ws = wb.create_sheet("Sensitivity Analysis")

    # Styles
    BLK = Font(name='Calibri', color='000000', size=11)
    BLK_B = Font(name='Calibri', color='000000', size=11, bold=True)
    WHT_B = Font(name='Calibri', color='FFFFFF', size=11, bold=True)
    GRN = Font(name='Calibri', color='008000', size=11)
    RED = Font(name='Calibri', color='CC0000', size=11)
    GRN_I = Font(name='Calibri', color='008000', size=11, italic=True)
    HDR = PatternFill('solid', fgColor='4472C4')
    INP = PatternFill('solid', fgColor='BDD7EE')
    GRN_BG = PatternFill('solid', fgColor='E2EFDA')
    ORANGE = PatternFill('solid', fgColor='FF6600')
    THIN = Border(
        left=Side('thin', color='B4C6E7'), right=Side('thin', color='B4C6E7'),
        top=Side('thin', color='B4C6E7'), bottom=Side('thin', color='B4C6E7'))
    CTR = Alignment(horizontal='center')

    def sc(r, c, v, f=BLK, fi=None, nf=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.alignment = CTR; cell.border = THIN
        if fi: cell.fill = fi
        if nf: cell.number_format = nf
        return cell

    def lb(r, c, v, f=BLK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = f; cell.border = THIN
        return cell

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 16
    for i in range(3, 15):
        ws.column_dimensions[cl(i)].width = 12

    mkt_price = cfg.get('stock_price', 0)
    base_growth = cfg['revenue_growth']
    base_margins = cfg['op_margins']
    wacc = cfg['_wacc']
    avg_margin = sum(base_margins) / len(base_margins)
    n_proj = len(base_growth)

    def cagr_from_list(rates):
        prod = 1
        for g in rates: prod *= (1 + g)
        return prod ** (1/len(rates)) - 1

    base_cagr = cagr_from_list(base_growth)

    def run_scenario_with_overrides(growth_rate=None, margin=None, wacc_override=None):
        """Run DCF with optional overrides. Uses base values where not overridden."""
        gr = [growth_rate] * n_proj if growth_rate is not None else list(base_growth)
        mg = [margin] * n_proj if margin is not None else list(base_margins)
        if wacc_override is not None:
            old_wacc = cfg['_wacc']
            cfg['_wacc'] = wacc_override
            price = run_dcf_fn(cfg, gr, mg)
            cfg['_wacc'] = old_wacc
        else:
            price = run_dcf_fn(cfg, gr, mg)
        return price

    def build_matrix(start_row, title, subtitle, row_label, col_label,
                     row_values, col_values, row_param, col_param):
        """Build a 2D sensitivity matrix. Returns end row."""
        r = start_row
        n_cols = len(col_values)

        # Header
        for c in range(2, 3 + n_cols + 1):
            ws.cell(row=r, column=c).fill = HDR
            ws.cell(row=r, column=c).font = WHT_B
            ws.cell(row=r, column=c).border = THIN
        ws.cell(row=r, column=2, value=title)

        r += 1
        ws.cell(row=r, column=2, value=subtitle)
        ws.cell(row=r, column=2).font = GRN_I

        # Column headers
        r += 1
        lb(r, 2, f"{row_label} \\ {col_label}", BLK_B)
        for j, cv in enumerate(col_values):
            fmt = '0.0%'
            sc(r, 3 + j, cv, BLK_B, None, fmt)

        # Find closest to market price
        closest = None
        closest_diff = float('inf')
        all_prices = {}

        for rv in row_values:
            for cv in col_values:
                kwargs = {}
                if row_param == 'growth': kwargs['growth_rate'] = rv
                elif row_param == 'margin': kwargs['margin'] = rv
                elif row_param == 'wacc': kwargs['wacc_override'] = rv

                if col_param == 'growth': kwargs['growth_rate'] = cv
                elif col_param == 'margin': kwargs['margin'] = cv
                elif col_param == 'wacc': kwargs['wacc_override'] = cv

                price = run_scenario_with_overrides(**kwargs)
                all_prices[(rv, cv)] = price
                diff = abs(price - mkt_price)
                if diff < closest_diff:
                    closest_diff = diff
                    closest = (rv, cv)

        # Fill matrix
        for rv in row_values:
            r += 1
            sc(r, 2, rv, BLK_B, None, '0.0%')
            for j, cv in enumerate(col_values):
                price = all_prices[(rv, cv)]
                is_closest = (rv, cv) == closest

                if is_closest:
                    cell_font = WHT_B
                    cell_fill = ORANGE
                elif price >= mkt_price:
                    cell_font = GRN
                    cell_fill = GRN_BG
                else:
                    cell_font = RED
                    cell_fill = None

                sc(r, 3 + j, round(price, 2), cell_font, cell_fill, '$#,##0')

        # Legend
        r += 2
        lb(r, 2, "Legenda", BLK_B)
        c = ws.cell(row=r, column=3, value="■")
        c.font = WHT_B; c.fill = ORANGE; c.border = THIN
        lb(r, 4, "Dichtst bij marktprijs (wat de markt impliceert)")

        r += 1
        c = ws.cell(row=r, column=3, value="■")
        c.font = GRN; c.fill = GRN_BG; c.border = THIN
        lb(r, 4, "Prijs boven markt → ondergewaardeerd")

        r += 1
        c = ws.cell(row=r, column=3, value="■")
        c.font = RED; c.border = THIN
        lb(r, 4, "Prijs onder markt → overgewaardeerd")

        return r + 1

    # ── Reverse DCF: find implied growth and margin at current price ──
    # Binary search for the flat growth rate that produces the market price
    def find_implied(param, lo, hi, tol=0.5, max_iter=40):
        """Binary search for the param value that matches market price."""
        for _ in range(max_iter):
            mid = (lo + hi) / 2
            if param == 'growth':
                price = run_scenario_with_overrides(growth_rate=mid)
            elif param == 'margin':
                price = run_scenario_with_overrides(margin=mid)
            else:
                return mid
            if abs(price - mkt_price) < tol:
                return mid
            if price > mkt_price:
                hi = mid
            else:
                lo = mid
        return (lo + hi) / 2

    implied_growth = find_implied('growth', -0.05, 0.50)
    implied_margin = find_implied('margin', 0.01, 0.80)

    # Define test ranges centered on implied values (reverse DCF midpoint)
    growth_step = 0.02
    growth_center = round(implied_growth / growth_step) * growth_step  # Snap to grid
    growth_tests = [round(growth_center + (i - 4) * growth_step, 4) for i in range(9)]
    growth_tests = [max(0.0, g) for g in growth_tests]

    margin_step = 0.02
    margin_center = round(implied_margin / margin_step) * margin_step
    margin_tests = [round(margin_center + (i - 4) * margin_step, 4) for i in range(9)]
    margin_tests = [max(0.05, m) for m in margin_tests]

    wacc_center = round(wacc, 3)
    wacc_tests = [round(wacc_center + (i - 4) * 0.005, 4) for i in range(9)]
    wacc_tests = [max(0.05, w) for w in wacc_tests]

    # Matrix 1: Revenue CAGR × Operating Margin
    end1 = build_matrix(
        2,
        "Matrix 1: Revenue CAGR × Operating Margin",
        f"Bij WACC = {wacc:.2%} | Marktprijs: ${mkt_price:.2f} | Implied CAGR: {implied_growth:.1%}, Implied Margin: {implied_margin:.1%}",
        "CAGR", "Marge",
        growth_tests, margin_tests,
        'growth', 'margin'
    )

    # Matrix 2: Revenue CAGR × WACC
    end2 = build_matrix(
        end1 + 3,
        "Matrix 2: Revenue CAGR × WACC",
        f"Bij Gem. Marge = {avg_margin:.1%} | Marktprijs: ${mkt_price:.2f} | Implied CAGR: {implied_growth:.1%}",
        "CAGR", "WACC",
        growth_tests, wacc_tests,
        'growth', 'wacc'
    )

    # Matrix 3: Operating Margin × WACC
    end3 = build_matrix(
        end2 + 3,
        "Matrix 3: Operating Margin × WACC",
        f"Bij Gem. CAGR = {base_cagr:.1%} | Marktprijs: ${mkt_price:.2f} | Implied Margin: {implied_margin:.1%}",
        "Marge", "WACC",
        margin_tests, wacc_tests,
        'margin', 'wacc'
    )

    # Overall synthesis
    r = end3 + 2
    lb(r, 2, "Synthese", BLK_B)
    r += 1
    ws.cell(row=r, column=2, value="Matrix 1 toont welke groei/marge-combinatie de marktprijs rechtvaardigt.")
    ws.cell(row=r, column=2).font = GRN_I
    r += 1
    ws.cell(row=r, column=2, value="Matrix 2 toont dat WACC typisch de grootste hefboom is — 1% WACC-verschil > 3% extra groei.")
    ws.cell(row=r, column=2).font = GRN_I
    r += 1
    ws.cell(row=r, column=2, value="Matrix 3 toont de interactie tussen marge en discount rate.")
    ws.cell(row=r, column=2).font = GRN_I

    return ws
